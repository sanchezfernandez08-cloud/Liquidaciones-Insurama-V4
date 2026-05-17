"""
Sistema de Liquidación de Siniestros — Insurama / A Plus Ajustadores
=====================================================================
Lee el Excel de siniestros Mapfre y genera informes Word de pago o rechazo,
reemplazando los marcadores XXXXXXXXXX con los datos de cada fila.
"""

import streamlit as st
import subprocess, sys

# Auto-instalar openpyxl si no está disponible (necesario en Streamlit Cloud)
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

import zipfile
import re
import io
from datetime import datetime, date

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Liquidación Siniestros — Insurama",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    .stApp { max-width: 1200px; margin: auto; }
    .block-container { padding-top: 1.5rem; }
    div[data-testid="stMetricValue"] { font-size: 1.6rem; }
    .tag-pago    { background:#d4edda; color:#155724; padding:3px 10px;
                   border-radius:12px; font-size:12px; font-weight:600; }
    .tag-rechazo { background:#fff3cd; color:#856404; padding:3px 10px;
                   border-radius:12px; font-size:12px; font-weight:600; }
    .tag-error   { background:#f8d7da; color:#721c24; padding:3px 10px;
                   border-radius:12px; font-size:12px; font-weight:600; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
MESES_ES = {
    1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",
    7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"
}

def fmt_fecha(val):
    """Formatea una fecha como '15 de abril de 2026'."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        d = val if isinstance(val, date) else val.date()
        return f"{d.day} de {MESES_ES[d.month]} de {d.year}"
    return str(val)

def fmt_fecha_corta(val):
    """Formatea como '15/04/2026'."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        d = val if isinstance(val, date) else val.date()
        return f"{d.day:02d}/{d.month:02d}/{d.year}"
    return str(val)

def fmt_dinero(val):
    """Formatea número como '$1.239.990'."""
    if val is None or val == "":
        return "$0"
    try:
        n = int(float(str(val).replace(".", "").replace(",", ".")))
        return f"${n:,.0f}".replace(",", ".")
    except:
        return str(val)

def limpiar_nombre(texto):
    """Corrige encoding latin1 mal interpretado como utf-8."""
    if texto is None:
        return ""
    try:
        # Intenta corregir doble-encodeo
        corregido = texto.encode("latin1").decode("utf-8")
        return corregido.strip()
    except:
        return str(texto).strip()

def limpiar_rut(val):
    """Normaliza el RUT a formato con puntos y guion."""
    if val is None:
        return ""
    s = str(val).strip()
    # Si es número puro (sin guion), agregar guion antes del último dígito
    if re.match(r'^\d{7,9}$', s.replace(".", "")):
        s = s.replace(".", "")
        return f"{s[:-1]}-{s[-1]}"
    return s

# ─────────────────────────────────────────────────────────────────────────────
# LECTURA DEL EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def leer_excel(archivo):
    """
    Lee el Excel de Mapfre y retorna lista de dicts normalizados.
    Columnas: ver encabezados en el archivo.
    """
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    filas = []
    for row in range(2, ws.max_row + 1):
        v = lambda c: ws.cell(row=row, column=c).value

        # Columnas del Excel (1-indexed):
        # 1=Alias siniestro, 2=Siniestro Mapfre, 3=Estado, 4=Fecha creación,
        # 5=Fecha siniestro, 6=Fecha comunicación, 7=Fecha resolución,
        # 8=Núm. Póliza, 9=Subestado.Descripcion, 10=NombreCompleto2,
        # 11=RUT, 12=Reserva, 13=Valoración, 14=UF, 15=Tipología,
        # 16=Descripción bien, 17=None, 18=Poliza.FechaEfecto,
        # 19=Suma Asegurada, 20=Remanente, 21=Cobertura afectada,
        # 22=Deducible, 23=Mes cierre, 24=Mes envio cia, 25=Estado final,
        # 26=Póliza Mapfre, 27=Referencia enviada

        alias      = v(1)
        if alias is None:
            continue  # fila vacía

        valoracion = v(13)
        try:
            val_num = int(float(valoracion)) if valoracion is not None else 0
        except:
            val_num = 0

        tipo = "pago" if val_num > 0 else "rechazo"

        nombre     = limpiar_nombre(v(10))
        subestado  = limpiar_nombre(v(9))
        rut        = limpiar_rut(v(11))
        poliza_num = str(v(8) or "").strip()
        poliza_mapfre = str(v(26) or "").strip()
        fecha_sin  = v(5)
        fecha_den  = v(6)   # comunicación = denuncia
        fecha_asig = v(4)   # creación = asignación
        fecha_res  = v(7)   # resolución = cierre
        suma_aseg  = v(19)
        bien       = limpiar_nombre(v(16)) or ""
        tipologia  = limpiar_nombre(v(15)) or ""

        filas.append({
            "_fila":        row,
            "_tipo":        tipo,
            "alias":        str(alias).strip(),
            "siniestro_mapfre": str(v(2) or "").strip(),
            "nombre":       nombre,
            "rut":          rut,
            "poliza_num":   poliza_num,
            "poliza_mapfre": poliza_mapfre,
            "subestado":    subestado,
            "valoracion":   val_num,
            "suma_asegurada": suma_aseg,
            "bien":         bien,
            "tipologia":    tipologia,
            "fecha_siniestro": fecha_sin,
            "fecha_denuncia":  fecha_den,
            "fecha_asignacion": fecha_asig,
            "fecha_resolucion": fecha_res,
            "fecha_efecto":  v(18),
        })
    return filas

def validar_fila(f):
    """Retorna lista de errores de validación (vacía = OK)."""
    errs = []
    if not f["nombre"]:
        errs.append("Sin nombre de asegurado")
    if not f["rut"]:
        errs.append("Sin RUT")
    if not f["alias"]:
        errs.append("Sin alias siniestro")
    return errs

# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DE WORD — reemplazo de marcadores en XML
# ─────────────────────────────────────────────────────────────────────────────
def reemplazar_en_xml(xml: str, replacements: list[tuple]) -> str:
    """
    Aplica una lista de (patron_regex, valor) al texto XML.
    Preserva todo el formato XML, solo cambia el contenido de <w:t>.
    """
    for patron, valor in replacements:
        xml = re.sub(patron, valor, xml, count=1)
    return xml

def generar_word_pago(fila: dict, plantilla_bytes: bytes) -> bytes:
    """
    Genera el Word de PAGO reemplazando los 21 marcadores X en el XML.
    Retorna los bytes del docx generado.
    """
    fecha_hoy = datetime.today()
    dia_hoy   = str(fecha_hoy.day)
    mes_hoy   = MESES_ES[fecha_hoy.month]

    # Los reemplazos se aplican en orden de aparición en el XML.
    # Cada tupla: (regex que captura el bloque X exacto, valor nuevo)
    # Usamos X{n} para ser precisos por longitud de marcador.
    replacements = [
        # [1]  Siniestro N° → Alias siniestro
        (r'<w:t>X{12}</w:t>', fila["alias"]),
        # [2]  Alias → igual Alias siniestro
        (r'<w:t>X{12}</w:t>', fila["alias"]),
        # [3]  Día (XX)
        (r'<w:t>X{2}</w:t>', dia_hoy),
        # [4]  Mes (XXXXXX)
        (r'<w:t>X{6}</w:t>', mes_hoy),
        # [5]  Certificado / Núm. Póliza (20 X)
        (r'<w:t>X{20}</w:t>', fila["poliza_num"]),
        # [6]  Monto Asegurado (19 X)
        (r'<w:t>X{19}</w:t>', fmt_dinero(fila["suma_asegurada"])),
        # [7]  Vigencia desde (8 X)
        (r'<w:t>X{8}</w:t>', fmt_fecha_corta(fila["fecha_efecto"])),
        # [8]  Vigencia hasta (8 X) — usamos fecha siniestro como proxy
        (r'<w:t>X{8}</w:t>', fmt_fecha_corta(fila["fecha_siniestro"])),
        # [9]  Asegurado nombre (22 X)
        (r'<w:t>X{22}</w:t>', fila["nombre"]),
        # [10] RUT (28 X)
        (r'<w:t>X{28}</w:t>', fila["rut"]),
        # [11] Fecha siniestro (9 X)
        (r'<w:t>X{9}</w:t>', fmt_fecha_corta(fila["fecha_siniestro"])),
        # [12] Fecha denuncia (9 X)
        (r'<w:t>X{9}</w:t>', fmt_fecha_corta(fila["fecha_denuncia"])),
        # [13] Fecha asignación (9 X)
        (r'<w:t>X{9}</w:t>', fmt_fecha_corta(fila["fecha_asignacion"])),
        # [14] Descripción bien / materia asegurada (9 X — primer bloque)
        (r'<w:t>X{9}</w:t>', fila["tipologia"]),
        # [15] Materia asegurada (9 X)
        (r'<w:t>X{9}</w:t>', fila["bien"]),
        # [16] Proveedor (10 X)
        (r'<w:t>X{10}</w:t>', "INSURAMA"),
        # [17] Valoración $ (9 X)
        (r'<w:t>X{9}</w:t>', fmt_dinero(fila["valoracion"])),
        # [18] Fecha cierre (10 X)
        (r'<w:t>X{10}</w:t>', fmt_fecha_corta(fila["fecha_resolucion"])),
        # [19] Pago recomendado $ (9 X) — mismo valor que valoración
        (r'<w:t>X{9}</w:t>', fmt_dinero(fila["valoracion"])),
        # [20] Nombre asegurado datos finales (19 X)
        (r'<w:t>X{19}</w:t>', fila["nombre"]),
        # [21] Correo electrónico (21 X)
        (r'<w:t>X{21}</w:t>', ""),
    ]

    return _aplicar_reemplazos(plantilla_bytes, replacements)


def generar_word_rechazo(fila: dict, plantilla_bytes: bytes) -> bytes:
    """
    Genera el Word de RECHAZO reemplazando los 18 marcadores X en el XML.
    """
    fecha_hoy = datetime.today()
    dia_hoy   = str(fecha_hoy.day)
    mes_hoy   = MESES_ES[fecha_hoy.month]

    replacements = [
        # [1]  Siniestro N° (11 X)
        (r'<w:t>X{11}</w:t>', fila["alias"]),
        # [2]  Alias (8 X)
        (r'<w:t>X{8}</w:t>', fila["alias"]),
        # [3]  Día (2 X)
        (r'<w:t>X{2}</w:t>', dia_hoy),
        # [4]  Mes (8 X)
        (r'<w:t>X{8}</w:t>', mes_hoy),
        # [5]  Póliza Mapfre (25 X)
        (r'<w:t>X{25}</w:t>', fila["poliza_mapfre"] or fila["poliza_num"]),
        # [6]  Certificado (25 X)
        (r'<w:t>X{25}</w:t>', fila["poliza_num"]),
        # [7]  Monto Asegurado $ (14 X)
        (r'<w:t>X{14}</w:t>', fmt_dinero(fila["suma_asegurada"])),
        # [8]  Vigencia desde (8 X)
        (r'<w:t>X{8}</w:t>', fmt_fecha_corta(fila["fecha_efecto"])),
        # [9]  Vigencia hasta (10 X)
        (r'<w:t>X{10}</w:t>', fmt_fecha_corta(fila["fecha_siniestro"])),
        # [10] Asegurado nombre (24 X)
        (r'<w:t>X{24}</w:t>', fila["nombre"]),
        # [11] RUT (31 X)
        (r'<w:t>X{31}</w:t>', fila["rut"]),
        # [12] Descripción bien (10 X)
        (r'<w:t>X{10}</w:t>', fila["bien"][:50] if fila["bien"] else fila["tipologia"]),
        # [13] Fecha siniestro (10 X)
        (r'<w:t>X{10}</w:t>', fmt_fecha_corta(fila["fecha_siniestro"])),
        # [14] Fecha denuncia (10 X)
        (r'<w:t>X{10}</w:t>', fmt_fecha_corta(fila["fecha_denuncia"])),
        # [15] Fecha asignación (10 X)
        (r'<w:t>X{10}</w:t>', fmt_fecha_corta(fila["fecha_asignacion"])),
        # [16] Descripción siniestro (9 X)
        (r'<w:t>X{9}</w:t>', fila["subestado"]),
        # [17] Nombre asegurado datos finales (22 X)
        (r'<w:t>X{22}</w:t>', fila["nombre"]),
        # [18] Correo electrónico (23 X)
        (r'<w:t>X{23}</w:t>', ""),
    ]

    return _aplicar_reemplazos(plantilla_bytes, replacements)


def _aplicar_reemplazos(plantilla_bytes: bytes, replacements: list) -> bytes:
    """
    Abre el docx como ZIP, modifica document.xml y retorna el nuevo docx.
    """
    entrada = io.BytesIO(plantilla_bytes)
    salida  = io.BytesIO()

    with zipfile.ZipFile(entrada, 'r') as zin:
        with zipfile.ZipFile(salida, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "word/document.xml":
                    xml = data.decode("utf-8")
                    for patron, valor in replacements:
                        # Escapar caracteres especiales XML en el valor
                        valor_xml = (str(valor)
                                     .replace("&", "&amp;")
                                     .replace("<", "&lt;")
                                     .replace(">", "&gt;")
                                     .replace('"', "&quot;"))
                        xml = re.sub(patron, f"<w:t>{valor_xml}</w:t>", xml, count=1)
                    data = xml.encode("utf-8")
                zout.writestr(item, data)

    salida.seek(0)
    return salida.read()


# ─────────────────────────────────────────────────────────────────────────────
# INTERFAZ STREAMLIT
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="background:#183363;color:white;padding:18px 24px;border-radius:8px;margin-bottom:1.2rem">
  <h2 style="margin:0;font-size:1.4rem">📋 Sistema de Liquidación de Siniestros</h2>
  <p style="margin:4px 0 0;opacity:.85;font-size:.9rem">Insurama / A Plus Ajustadores — Mapfre Cía. de Seguros Generales de Chile S.A.</p>
</div>
""", unsafe_allow_html=True)

# ── Panel de carga ──────────────────────────────────────────────────────────
with st.expander("📁 Cargar archivos", expanded=True):
    c1, c2, c3 = st.columns(3)
    with c1:
        excel_file = st.file_uploader(
            "Excel de siniestros Mapfre",
            type=["xlsx", "xls"],
            help="Archivo entregado por Mapfre con el cierre del período",
        )
    with c2:
        tmpl_pago = st.file_uploader(
            "Plantilla Word — PAGO / Indemnización",
            type=["docx"],
            help="Informe_de_Liquidación_Tipo_Indemnización.docx",
        )
    with c3:
        tmpl_rechazo = st.file_uploader(
            "Plantilla Word — RECHAZO",
            type=["docx"],
            help="Informe_de_Liquidación_Tipo_Rechazo.docx",
        )

# ── Validación de archivos ──────────────────────────────────────────────────
if not excel_file:
    st.info("📂 Cargue el Excel de siniestros Mapfre para comenzar.")
    st.stop()

if not tmpl_pago or not tmpl_rechazo:
    st.warning("⚠️ Cargue también ambas plantillas Word para poder generar los informes.")

# ── Lectura del Excel ────────────────────────────────────────────────────────
with st.spinner("Leyendo Excel de siniestros…"):
    try:
        filas = leer_excel(excel_file)
    except Exception as e:
        st.error(f"❌ Error al leer el Excel: {e}")
        st.stop()

n_pago    = sum(1 for f in filas if f["_tipo"] == "pago")
n_rechazo = sum(1 for f in filas if f["_tipo"] == "rechazo")
n_error   = sum(1 for f in filas if validar_fila(f))

# ── Métricas ─────────────────────────────────────────────────────────────────
m1, m2, m3, m4 = st.columns(4)
m1.metric("Total siniestros", len(filas))
m2.metric("✅ Pago",    n_pago)
m3.metric("⚠️ Rechazo", n_rechazo)
m4.metric("❌ Con errores", n_error)

st.divider()

# ── Filtros ───────────────────────────────────────────────────────────────────
col_f1, col_f2, col_f3 = st.columns([2, 2, 3])
with col_f1:
    filtro_tipo = st.selectbox(
        "Filtrar por tipo",
        ["Todos", "Solo PAGO", "Solo RECHAZO"],
    )
with col_f2:
    filtro_estado = st.selectbox(
        "Filtrar por estado",
        ["Todos", "Sin errores", "Con errores"],
    )
with col_f3:
    buscar = st.text_input("🔍 Buscar por nombre, alias o RUT", "")

# Aplicar filtros
filas_vis = filas[:]
if filtro_tipo == "Solo PAGO":
    filas_vis = [f for f in filas_vis if f["_tipo"] == "pago"]
elif filtro_tipo == "Solo RECHAZO":
    filas_vis = [f for f in filas_vis if f["_tipo"] == "rechazo"]
if filtro_estado == "Sin errores":
    filas_vis = [f for f in filas_vis if not validar_fila(f)]
elif filtro_estado == "Con errores":
    filas_vis = [f for f in filas_vis if validar_fila(f)]
if buscar.strip():
    q = buscar.strip().lower()
    filas_vis = [f for f in filas_vis
                 if q in f["nombre"].lower()
                 or q in f["alias"].lower()
                 or q in f["rut"].lower()]

st.caption(f"Mostrando {len(filas_vis)} de {len(filas)} registros")

# ── Generación masiva ─────────────────────────────────────────────────────────
if tmpl_pago and tmpl_rechazo:
    st.markdown("### ⚙️ Generación de informes")
    gcol1, gcol2, gcol3 = st.columns(3)

    with gcol1:
        gen_filtrados = st.button(
            f"📝 Generar informes filtrados ({len(filas_vis)})",
            type="primary",
            use_container_width=True,
            disabled=len(filas_vis) == 0,
        )
    with gcol2:
        gen_pago = st.button(
            f"✅ Generar solo PAGO ({n_pago})",
            use_container_width=True,
        )
    with gcol3:
        gen_rechazo = st.button(
            f"⚠️ Generar solo RECHAZO ({n_rechazo})",
            use_container_width=True,
        )

    # Determinar qué generar
    filas_a_generar = []
    if gen_filtrados:
        filas_a_generar = [f for f in filas_vis if not validar_fila(f)]
    elif gen_pago:
        filas_a_generar = [f for f in filas if f["_tipo"] == "pago" and not validar_fila(f)]
    elif gen_rechazo:
        filas_a_generar = [f for f in filas if f["_tipo"] == "rechazo" and not validar_fila(f)]

    if filas_a_generar:
        bytes_pago    = tmpl_pago.read()    if tmpl_pago    else None
        bytes_rechazo = tmpl_rechazo.read() if tmpl_rechazo else None

        # Restablecer posición para relecturas
        if tmpl_pago:    tmpl_pago.seek(0)
        if tmpl_rechazo: tmpl_rechazo.seek(0)

        progreso = st.progress(0, text="Generando informes…")
        docs_ok   = []
        docs_err  = []

        for i, fila in enumerate(filas_a_generar):
            try:
                if fila["_tipo"] == "pago":
                    doc_bytes = generar_word_pago(fila, bytes_pago)
                else:
                    doc_bytes = generar_word_rechazo(fila, bytes_rechazo)

                nombre_archivo = (
                    f"Informe_{fila['alias']}_{fila['_tipo'].upper()}_"
                    f"{fila['nombre'].replace(' ','_')[:30]}.docx"
                )
                docs_ok.append((nombre_archivo, doc_bytes, fila))

            except Exception as e:
                docs_err.append((fila["alias"], fila["nombre"], str(e)))

            progreso.progress((i + 1) / len(filas_a_generar),
                              text=f"Generando {i+1}/{len(filas_a_generar)}: {fila['alias']}")

        progreso.empty()

        # Resultados
        st.success(f"✅ {len(docs_ok)} informes generados · ❌ {len(docs_err)} errores")

        if docs_err:
            with st.expander("❌ Ver errores"):
                for alias, nombre, err in docs_err:
                    st.error(f"**{alias}** — {nombre}: {err}")

        # Descargas individuales
        if docs_ok:
            st.markdown("#### 📥 Descargar informes")
            for nombre_archivo, doc_bytes, fila in docs_ok:
                tipo_tag = "pago" if fila["_tipo"] == "pago" else "rechazo"
                tipo_lbl = "✅ PAGO" if fila["_tipo"] == "pago" else "⚠️ RECHAZO"
                col_a, col_b, col_c = st.columns([3, 2, 2])
                with col_a:
                    st.markdown(f"**{fila['alias']}** — {fila['nombre']}")
                with col_b:
                    st.markdown(f"<span class='tag-{tipo_tag}'>{tipo_lbl}</span>",
                                unsafe_allow_html=True)
                with col_c:
                    st.download_button(
                        "⬇️ Descargar",
                        data=doc_bytes,
                        file_name=nombre_archivo,
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        key=f"dl_{fila['alias']}",
                        use_container_width=True,
                    )

            # ZIP con todos
            st.divider()
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for nombre_archivo, doc_bytes, _ in docs_ok:
                    zf.writestr(nombre_archivo, doc_bytes)
            zip_buf.seek(0)
            st.download_button(
                f"📦 Descargar todos ({len(docs_ok)}) en ZIP",
                data=zip_buf.read(),
                file_name=f"Informes_Liquidacion_{datetime.today().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
            )

st.divider()

# ── Tabla de registros ────────────────────────────────────────────────────────
st.markdown("### 📋 Registros cargados")

for fila in filas_vis:
    errores = validar_fila(fila)
    tipo    = fila["_tipo"]

    # Color de fondo según tipo
    if errores:
        bg = "#fff0f0"
        tag = "<span class='tag-error'>❌ ERROR</span>"
    elif tipo == "pago":
        bg = "#f0fff4"
        tag = "<span class='tag-pago'>✅ PAGO</span>"
    else:
        bg = "#fffdf0"
        tag = "<span class='tag-rechazo'>⚠️ RECHAZO</span>"

    with st.expander(
        f"{fila['alias']}  —  {fila['nombre']}  |  {fila['subestado']}",
        expanded=False,
    ):
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"**Tipo:** {tag}", unsafe_allow_html=True)
            st.markdown(f"**Alias siniestro:** {fila['alias']}")
            st.markdown(f"**Siniestro Mapfre:** {fila['siniestro_mapfre']}")
            st.markdown(f"**Subestado:** {fila['subestado']}")
        with col2:
            st.markdown(f"**Nombre asegurado:** {fila['nombre']}")
            st.markdown(f"**RUT:** {fila['rut']}")
            st.markdown(f"**Póliza (Núm.):** {fila['poliza_num']}")
            st.markdown(f"**Póliza Mapfre:** {fila['poliza_mapfre']}")
        with col3:
            st.markdown(f"**Valoración:** {fmt_dinero(fila['valoracion'])}")
            st.markdown(f"**Suma asegurada:** {fmt_dinero(fila['suma_asegurada'])}")
            st.markdown(f"**Fecha siniestro:** {fmt_fecha(fila['fecha_siniestro'])}")
            st.markdown(f"**Fecha resolución:** {fmt_fecha(fila['fecha_resolucion'])}")
            st.markdown(f"**Bien asegurado:** {fila['bien'][:60]}")

        if errores:
            for e in errores:
                st.error(f"⚠️ {e}")

        # Botón descarga individual (solo si las plantillas están cargadas)
        if tmpl_pago and tmpl_rechazo and not errores:
            if tipo == "pago":
                tmpl_pago.seek(0)
                b = generar_word_pago(fila, tmpl_pago.read())
                tmpl_pago.seek(0)
            else:
                tmpl_rechazo.seek(0)
                b = generar_word_rechazo(fila, tmpl_rechazo.read())
                tmpl_rechazo.seek(0)

            st.download_button(
                f"📄 Descargar informe — {fila['alias']}",
                data=b,
                file_name=f"Informe_{fila['alias']}_{tipo.upper()}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                key=f"ind_{fila['alias']}_{fila['_fila']}",
            )
